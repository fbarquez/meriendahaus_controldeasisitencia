"""
Django Admin configuration for the time tracking system.
Enhanced with dashboard, reports, and better UX.
"""

import csv
from datetime import timedelta
from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.utils import timezone
from django.utils.html import format_html
from django.urls import reverse
from django import forms

from simple_history.admin import SimpleHistoryAdmin

from .models import Location, TimeEntry, FailedClockAttempt


# -----------------------------------------------------------------------------
# Custom Admin Site
# -----------------------------------------------------------------------------

class ZeiterfassungAdminSite(admin.AdminSite):
    site_header = 'Meriendahaus - Registro Horario'
    site_title = 'Zeiterfassung'
    index_title = 'Panel de Administración'

    def index(self, request, extra_context=None):
        """Redirect admin index to dashboard."""
        from django.shortcuts import redirect
        return redirect('admin_dashboard')


# Replace default admin site
admin.site = ZeiterfassungAdminSite(name='admin')


# -----------------------------------------------------------------------------
# Location Admin with QR generation
# -----------------------------------------------------------------------------

@admin.register(Location, site=admin.site)
class LocationAdmin(SimpleHistoryAdmin):
    list_display = ('code', 'name', 'is_active', 'get_ips_display', 'get_qr_link', 'updated_at')
    list_filter = ('is_active',)
    search_fields = ('code', 'name')
    readonly_fields = ('created_at', 'updated_at', 'get_qr_preview')

    fieldsets = (
        ('Información', {
            'fields': ('code', 'name', 'is_active')
        }),
        ('Configuración de IP (Anti-fraude)', {
            'fields': ('allowed_ips',),
            'description': format_html(
                'Lista de IPs públicas permitidas para fichar. '
                'Formato JSON: ["85.123.45.67"]. '
                '<a href="https://whatismyip.com" target="_blank">Ver tu IP actual</a>'
            )
        }),
        ('Código QR', {
            'fields': ('get_qr_preview',),
            'description': 'Código QR para imprimir y colocar en el local.'
        }),
        ('Metadatos', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def get_ips_display(self, obj):
        ips = obj.allowed_ips or []
        if not ips:
            return format_html('<span style="color: #e53e3e; font-weight: 500;">Sin configurar</span>')
        display = ', '.join(str(ip) for ip in ips[:2])
        if len(ips) > 2:
            display += f' (+{len(ips) - 2})'
        return format_html('<code style="background: #edf2f7; padding: 2px 6px; border-radius: 3px;">{}</code>', display)
    get_ips_display.short_description = 'IPs Permitidas'

    def get_qr_link(self, obj):
        url = reverse('admin_qr_print', args=[obj.pk])
        return format_html(
            '<a href="{}" target="_blank" style="color: #3182ce; font-weight: 500;">Ver QR</a>',
            url
        )
    get_qr_link.short_description = 'Código QR'

    def get_qr_preview(self, obj):
        if not obj.pk:
            return "Guarda primero para generar el QR"
        qr_url = reverse('admin_generate_qr', args=[obj.pk])
        print_url = reverse('admin_qr_print', args=[obj.pk])
        return format_html(
            '<div style="text-align: center;">'
            '<img src="{}" style="width: 150px; height: 150px; border: 1px solid #e2e8f0; border-radius: 4px;"><br><br>'
            '<a href="{}" target="_blank" class="button">Imprimir QR</a>'
            '</div>',
            qr_url, print_url
        )
    get_qr_preview.short_description = 'Vista previa QR'


# -----------------------------------------------------------------------------
# TimeEntry Admin with enhanced features
# -----------------------------------------------------------------------------

class TimeEntryForm(forms.ModelForm):
    """Custom form that requires notes for manual entries."""

    class Meta:
        model = TimeEntry
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        is_manual = cleaned_data.get('is_manual', False)
        notes = cleaned_data.get('notes', '').strip()

        if self.instance.pk or is_manual:
            if len(notes) < 10:
                raise forms.ValidationError(
                    'Las notas son obligatorias para entradas manuales '
                    '(mínimo 10 caracteres explicando el motivo)'
                )

        return cleaned_data


class WeekFilter(admin.SimpleListFilter):
    title = 'Semana'
    parameter_name = 'week'

    def lookups(self, request, model_admin):
        return [
            ('current', 'Esta semana'),
            ('last', 'Semana pasada'),
        ]

    def queryset(self, request, queryset):
        today = timezone.now().date()
        start_of_week = today - timedelta(days=today.weekday())

        if self.value() == 'current':
            return queryset.filter(check_in__date__gte=start_of_week)
        elif self.value() == 'last':
            last_week_start = start_of_week - timedelta(days=7)
            return queryset.filter(
                check_in__date__gte=last_week_start,
                check_in__date__lt=start_of_week
            )
        return queryset


class StatusFilter(admin.SimpleListFilter):
    title = 'Estado'
    parameter_name = 'status'

    def lookups(self, request, model_admin):
        return [
            ('open', 'Fichajes abiertos'),
            ('closed', 'Fichajes cerrados'),
            ('overtime', 'Más de 8 horas'),
            ('manual', 'Entradas manuales'),
        ]

    def queryset(self, request, queryset):
        if self.value() == 'open':
            return queryset.filter(check_out__isnull=True)
        elif self.value() == 'closed':
            return queryset.filter(check_out__isnull=False)
        elif self.value() == 'overtime':
            eight_hours = timedelta(hours=8)
            return queryset.filter(check_out__isnull=False).extra(
                where=["check_out - check_in > %s"],
                params=[eight_hours]
            )
        elif self.value() == 'manual':
            return queryset.filter(is_manual=True)
        return queryset


class DateRangeFilter(admin.SimpleListFilter):
    title = 'Período'
    parameter_name = 'period'

    def lookups(self, request, model_admin):
        return [
            ('today', 'Hoy'),
            ('yesterday', 'Ayer'),
            ('week', 'Esta semana'),
            ('month', 'Este mes'),
            ('last_month', 'Mes anterior'),
        ]

    def queryset(self, request, queryset):
        today = timezone.now().date()

        if self.value() == 'today':
            return queryset.filter(check_in__date=today)
        elif self.value() == 'yesterday':
            return queryset.filter(check_in__date=today - timedelta(days=1))
        elif self.value() == 'week':
            start = today - timedelta(days=today.weekday())
            return queryset.filter(check_in__date__gte=start)
        elif self.value() == 'month':
            return queryset.filter(
                check_in__year=today.year,
                check_in__month=today.month
            )
        elif self.value() == 'last_month':
            first_of_month = today.replace(day=1)
            last_month = first_of_month - timedelta(days=1)
            return queryset.filter(
                check_in__year=last_month.year,
                check_in__month=last_month.month
            )
        return queryset


@admin.register(TimeEntry, site=admin.site)
class TimeEntryAdmin(SimpleHistoryAdmin):
    form = TimeEntryForm
    list_display = (
        'get_employee',
        'get_date',
        'get_check_in_time',
        'get_check_out_time',
        'get_duration',
        'location',
        'get_status_badge'
    )
    list_filter = (
        DateRangeFilter,
        StatusFilter,
        'location',
        'user',
        'is_manual',
    )
    search_fields = (
        'user__username',
        'user__first_name',
        'user__last_name',
        'notes'
    )
    date_hierarchy = 'check_in'
    raw_id_fields = ('user',)
    readonly_fields = (
        'check_in_ip',
        'check_out_ip',
        'created_at',
        'modified_at',
        'modified_by'
    )
    list_per_page = 50

    fieldsets = (
        ('Datos del Fichaje', {
            'fields': ('user', 'location', 'check_in', 'check_out')
        }),
        ('Entrada Manual / Corrección', {
            'fields': ('is_manual', 'notes'),
            'description': (
                '⚠️ Marcar como "Entrada manual" y añadir notas explicativas '
                'si estás corrigiendo o creando un fichaje manualmente.'
            )
        }),
        ('Información Técnica', {
            'fields': ('check_in_ip', 'check_out_ip'),
            'classes': ('collapse',)
        }),
        ('Metadatos', {
            'fields': ('created_at', 'modified_at', 'modified_by'),
            'classes': ('collapse',)
        }),
    )

    actions = ['export_csv', 'export_excel', 'mark_as_closed']

    class Media:
        css = {
            'all': ('admin/css/custom.css',)
        }

    def get_employee(self, obj):
        name = obj.user.get_full_name() or obj.user.username
        initials = ''.join([n[0].upper() for n in name.split()[:2]]) if name else '?'
        return format_html(
            '<span style="display: inline-flex; align-items: center; gap: 8px;">'
            '<span style="background: #3498db; color: white; width: 28px; height: 28px; '
            'border-radius: 50%; display: inline-flex; align-items: center; justify-content: center; '
            'font-size: 11px; font-weight: 600;">{}</span>'
            '{}</span>',
            initials, name
        )
    get_employee.short_description = 'Empleado'
    get_employee.admin_order_field = 'user__first_name'

    def get_date(self, obj):
        return obj.check_in.strftime('%d/%m/%Y')
    get_date.short_description = 'Fecha'
    get_date.admin_order_field = 'check_in'

    def get_check_in_time(self, obj):
        return format_html(
            '<span style="font-family: monospace; font-weight: 600;">{}</span>',
            obj.check_in.strftime('%H:%M')
        )
    get_check_in_time.short_description = 'Entrada'

    def get_check_out_time(self, obj):
        if obj.check_out:
            return format_html(
                '<span style="font-family: monospace; font-weight: 600;">{}</span>',
                obj.check_out.strftime('%H:%M')
            )
        return format_html('<span style="color: #f39c12;">—</span>')
    get_check_out_time.short_description = 'Salida'

    def get_duration(self, obj):
        display = obj.duration_display
        if obj.duration_minutes and obj.duration_minutes > 480:
            return format_html(
                '<span style="color: #e67e22; font-weight: 600;">{}</span>',
                display
            )
        return display
    get_duration.short_description = 'Duración'

    def get_status_badge(self, obj):
        if obj.is_manual:
            return format_html(
                '<span style="background: #fefcbf; color: #975a16; '
                'padding: 4px 10px; border-radius: 4px; font-size: 11px; font-weight: 600; '
                'text-transform: uppercase; letter-spacing: 0.03em;">Manual</span>'
            )
        if obj.is_open:
            return format_html(
                '<span style="background: #bee3f8; color: #2c5282; '
                'padding: 4px 10px; border-radius: 4px; font-size: 11px; font-weight: 600; '
                'text-transform: uppercase; letter-spacing: 0.03em;">Abierto</span>'
            )
        if obj.duration_minutes and obj.duration_minutes > 480:
            return format_html(
                '<span style="background: #fed7d7; color: #c53030; '
                'padding: 4px 10px; border-radius: 4px; font-size: 11px; font-weight: 600; '
                'text-transform: uppercase; letter-spacing: 0.03em;">+8h</span>'
            )
        return format_html(
            '<span style="background: #c6f6d5; color: #276749; '
            'padding: 4px 10px; border-radius: 4px; font-size: 11px; font-weight: 600; '
            'text-transform: uppercase; letter-spacing: 0.03em;">OK</span>'
        )
    get_status_badge.short_description = 'Estado'

    def save_model(self, request, obj, form, change):
        if change:
            obj.modified_by = request.user
            obj.is_manual = True
        else:
            obj.is_manual = True
            obj.check_in_ip = 'ADMIN'
            if obj.check_out:
                obj.check_out_ip = 'ADMIN'

        super().save_model(request, obj, form, change)

    @admin.action(description='Exportar a CSV')
    def export_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = (
            f'attachment; filename="fichajes_{timezone.now().strftime("%Y%m%d")}.csv"'
        )
        response.write('\ufeff')

        writer = csv.writer(response, delimiter=';')
        writer.writerow([
            'Empleado', 'Fecha', 'Entrada', 'Salida',
            'Duración', 'Local', 'Manual', 'Notas'
        ])

        for entry in queryset.select_related('user', 'location'):
            writer.writerow([
                entry.user.get_full_name() or entry.user.username,
                entry.check_in.strftime('%d/%m/%Y'),
                entry.check_in.strftime('%H:%M'),
                entry.check_out.strftime('%H:%M') if entry.check_out else '',
                entry.duration_display,
                entry.location.name,
                'Sí' if entry.is_manual else 'No',
                entry.notes or ''
            ])

        return response

    @admin.action(description='Exportar a Excel')
    def export_excel(self, request, queryset):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.message_user(request, 'openpyxl no instalado.', level='error')
            return

        wb = Workbook()
        ws = wb.active
        ws.title = 'Fichajes'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')

        headers = ['Empleado', 'Fecha', 'Entrada', 'Salida', 'Duración (min)', 'Local', 'Manual', 'Notas']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row, entry in enumerate(queryset.select_related('user', 'location'), 2):
            ws.cell(row=row, column=1, value=entry.user.get_full_name() or entry.user.username)
            ws.cell(row=row, column=2, value=entry.check_in.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=3, value=entry.check_in.strftime('%H:%M'))
            ws.cell(row=row, column=4, value=entry.check_out.strftime('%H:%M') if entry.check_out else '')
            ws.cell(row=row, column=5, value=entry.duration_minutes or 0)
            ws.cell(row=row, column=6, value=entry.location.name)
            ws.cell(row=row, column=7, value='Sí' if entry.is_manual else 'No')
            ws.cell(row=row, column=8, value=entry.notes or '')

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="fichajes_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        )
        wb.save(response)
        return response

    @admin.action(description='Cerrar fichajes seleccionados')
    def mark_as_closed(self, request, queryset):
        count = 0
        for entry in queryset.filter(check_out__isnull=True):
            end_of_day = timezone.make_aware(
                timezone.datetime.combine(entry.check_in.date(), timezone.datetime.max.time())
            )
            entry.check_out = end_of_day
            entry.check_out_ip = 'ADMIN_CLOSED'
            entry.is_manual = True
            entry.notes = (entry.notes or '') + f'\n[Cerrado por admin {timezone.now().strftime("%d/%m/%Y %H:%M")}]'
            entry.modified_by = request.user
            entry.save()
            count += 1

        self.message_user(request, f'Se cerraron {count} fichajes.')


# -----------------------------------------------------------------------------
# Custom User Admin
# -----------------------------------------------------------------------------

class CustomUserAdmin(BaseUserAdmin):
    list_display = ('username', 'get_full_name_display', 'email', 'is_active', 'get_status')
    list_filter = ('is_active', 'is_staff')

    def get_full_name_display(self, obj):
        name = obj.get_full_name()
        if name:
            return name
        return format_html('<span style="color: #999;">Sin nombre</span>')
    get_full_name_display.short_description = 'Nombre'

    def get_status(self, obj):
        # Check if user has open entry
        if TimeEntry.objects.filter(user=obj, check_out__isnull=True).exists():
            return format_html(
                '<span style="background: #c6f6d5; color: #276749; padding: 3px 8px; '
                'border-radius: 4px; font-size: 11px; font-weight: 600;">PRESENTE</span>'
            )
        return format_html('<span style="color: #a0aec0;">—</span>')
    get_status.short_description = 'Estado'


admin.site.register(User, CustomUserAdmin)


# -----------------------------------------------------------------------------
# Failed Clock Attempts Admin
# -----------------------------------------------------------------------------

@admin.register(FailedClockAttempt, site=admin.site)
class FailedClockAttemptAdmin(admin.ModelAdmin):
    list_display = (
        'get_employee',
        'get_action_badge',
        'ip_address',
        'location',
        'timestamp'
    )
    list_filter = ('action', 'location', 'timestamp')
    search_fields = ('user__username', 'user__first_name', 'user__last_name', 'ip_address')
    date_hierarchy = 'timestamp'
    readonly_fields = ('user', 'location', 'action', 'ip_address', 'timestamp')
    list_per_page = 50

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def get_employee(self, obj):
        name = obj.user.get_full_name() or obj.user.username
        return format_html(
            '<span style="font-weight: 600;">{}</span>',
            name
        )
    get_employee.short_description = 'Employee'
    get_employee.admin_order_field = 'user__first_name'

    def get_action_badge(self, obj):
        if obj.action == 'in':
            return format_html(
                '<span style="background: #fed7d7; color: #c53030; '
                'padding: 4px 10px; border-radius: 4px; font-size: 11px; font-weight: 600; '
                'text-transform: uppercase;">Failed IN</span>'
            )
        return format_html(
            '<span style="background: #fed7d7; color: #c53030; '
            'padding: 4px 10px; border-radius: 4px; font-size: 11px; font-weight: 600; '
            'text-transform: uppercase;">Failed OUT</span>'
        )
    get_action_badge.short_description = 'Attempt'
